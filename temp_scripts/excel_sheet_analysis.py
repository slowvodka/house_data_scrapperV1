"""Analyze Excel sheets and document their purpose."""
import pandas as pd
from pathlib import Path
import openpyxl

excel_file = Path("דירה להשקעה עדכון  V3.2 8.24.xlsx")

wb = openpyxl.load_workbook(excel_file, data_only=False)
sheet_names = wb.sheetnames

# Read each sheet and understand its structure
analysis = {}

for sheet_name in sheet_names:
    print(f"\n{'='*80}")
    print(f"Analyzing: {sheet_name}")
    print('='*80)
    
    df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl', header=0)
    
    # Get column names
    cols = list(df.columns)
    
    # Sample data
    sample = df.head(3)
    
    # Check for formulas
    ws = wb[sheet_name]
    formulas = []
    for row in ws.iter_rows(min_row=2, max_row=min(5, ws.max_row)):
        for cell in row:
            if cell.data_type == 'f':
                formulas.append(f"{cell.coordinate}: {str(cell.value)[:100]}")
                if len(formulas) >= 5:
                    break
        if len(formulas) >= 5:
            break
    
    analysis[sheet_name] = {
        'rows': len(df),
        'cols': len(df.columns),
        'column_names': cols,
        'sample_data': sample.to_dict('records')[:2],
        'formulas': formulas[:5]
    }
    
    print(f"Rows: {len(df)}, Columns: {len(df.columns)}")
    print(f"Columns: {cols[:10]}...")  # First 10 columns

# Save analysis to markdown
with open("temp_scripts/excel_analysis_report.md", "w", encoding="utf-8") as f:
    f.write("# Excel File Analysis Report\n\n")
    f.write(f"File: {excel_file.name}\n")
    f.write(f"Total Sheets: {len(sheet_names)}\n\n")
    
    for sheet_name, data in analysis.items():
        f.write(f"## Sheet: {sheet_name}\n\n")
        f.write(f"- **Dimensions:** {data['rows']} rows × {data['cols']} columns\n")
        f.write(f"- **Purpose:** (See analysis below)\n\n")
        f.write("### Column Names:\n")
        for i, col in enumerate(data['column_names'], 1):
            f.write(f"{i}. {col}\n")
        f.write("\n### Sample Formulas:\n")
        for formula in data['formulas']:
            f.write(f"- {formula}\n")
        f.write("\n---\n\n")

print("\nAnalysis saved to temp_scripts/excel_analysis_report.md")

