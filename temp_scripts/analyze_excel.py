"""Analyze Excel file structure and content.

Reads the Excel file and extracts:
1. Sheet names
2. Column headers for each sheet
3. Sample data from each sheet
4. Formulas (if any)
5. Data types and structure
"""
import pandas as pd
from pathlib import Path
import openpyxl

# Excel file path
excel_file = Path("דירה להשקעה עדכון  V3.2 8.24.xlsx")

if not excel_file.exists():
    print(f"ERROR: File not found: {excel_file}")
    exit(1)

print("=" * 80)
print("EXCEL FILE ANALYSIS")
print("=" * 80)
print(f"File: {excel_file}")
print(f"Size: {excel_file.stat().st_size / 1024:.2f} KB")
print()

# Load workbook with openpyxl to get formulas
wb = openpyxl.load_workbook(excel_file, data_only=False)
sheet_names = wb.sheetnames

print(f"Total Sheets: {len(sheet_names)}")
print(f"Sheet Names: {', '.join(sheet_names)}")
print()

# Analyze each sheet
for idx, sheet_name in enumerate(sheet_names, 1):
    print("=" * 80)
    print(f"SHEET {idx}/{len(sheet_names)}: {sheet_name}")
    print("=" * 80)
    
    # Read with pandas for data analysis
    try:
        df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl')
        
        print(f"Dimensions: {df.shape[0]} rows × {df.shape[1]} columns")
        print()
        
        # Show column names
        print("Columns:")
        for col_idx, col_name in enumerate(df.columns, 1):
            non_null = df[col_name].notna().sum()
            dtype = str(df[col_name].dtype)
            print(f"  {col_idx:2d}. {col_name} ({dtype}, {non_null} non-null)")
        print()
        
        # Show sample data (first few rows)
        print("Sample Data (first 3 rows):")
        print(df.head(3).to_string())
        print()
        
        # Check for formulas in this sheet
        ws = wb[sheet_name]
        formula_count = 0
        formulas_found = []
        
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == 'f':  # Formula
                    formula_count += 1
                    if len(formulas_found) < 5:  # Show first 5 formulas
                        formulas_found.append({
                            'cell': cell.coordinate,
                            'formula': cell.value,
                            'value': cell.value if hasattr(cell, 'value') else None
                        })
        
        if formula_count > 0:
            print(f"Formulas Found: {formula_count}")
            print("Sample Formulas:")
            for f in formulas_found[:5]:
                print(f"  {f['cell']}: {f['formula']}")
            print()
        
        # Data summary
        print("Data Summary:")
        print(df.describe(include='all').to_string())
        print()
        
    except Exception as e:
        print(f"ERROR reading sheet: {e}")
        print()
    
    print()

print("=" * 80)
print("ANALYSIS COMPLETE")
print("=" * 80)

