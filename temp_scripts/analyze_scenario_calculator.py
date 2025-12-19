"""Analyze Scenario Calculator sheet in detail - identify inputs, formulas, and logic."""
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill

excel_file = Path("דירה להשקעה עדכון  V3.2 8.24.xlsx")

print("=" * 100)
print("SCENARIO CALCULATOR DETAILED ANALYSIS")
print("=" * 100)
print()

wb = openpyxl.load_workbook(excel_file, data_only=False)
ws = wb["Scenarios Calculator"]

# Read the sheet
df = pd.read_excel(excel_file, sheet_name="Scenarios Calculator", engine='openpyxl', header=None)

print("Sheet Dimensions:", df.shape)
print()

# Analyze cell colors (yellow = user input)
print("=" * 100)
print("IDENTIFYING USER INPUTS (Yellow Cells)")
print("=" * 100)

yellow_inputs = []
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row), start=1):
    for col_idx, cell in enumerate(row, start=1):
        if cell.fill and cell.fill.patternType == 'solid':
            fill_color = cell.fill.fgColor
            # Check if it's yellow (light yellow typically RGB around 255, 255, 200-255)
            if fill_color and fill_color.rgb:
                rgb = fill_color.rgb
                # Light yellow detection
                if 'FFFF' in rgb.upper() or 'FFE' in rgb.upper():
                    yellow_inputs.append({
                        'cell': cell.coordinate,
                        'row': row_idx,
                        'col': col_idx,
                        'value': cell.value,
                        'formula': cell.value if cell.data_type == 'f' else None
                    })

print(f"Found {len(yellow_inputs)} yellow (user input) cells:")
for inp in yellow_inputs[:20]:  # Show first 20
    print(f"  {inp['cell']}: {inp['value']}")

print()

# Read structured data
print("=" * 100)
print("SHEET STRUCTURE ANALYSIS")
print("=" * 100)

# Read with proper headers
df_header = pd.read_excel(excel_file, sheet_name="Scenarios Calculator", engine='openpyxl', header=0)

print("\nColumn Headers:")
for i, col in enumerate(df_header.columns, 1):
    print(f"  {i}. {col}")

print("\nFirst 30 rows:")
print(df_header.head(30).to_string())

# Analyze formulas in detail
print("\n" + "=" * 100)
print("FORMULA ANALYSIS")
print("=" * 100)

formulas_by_row = {}
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(50, ws.max_row)), start=1):
    row_formulas = []
    for col_idx, cell in enumerate(row, start=1):
        if cell.data_type == 'f':
            row_formulas.append({
                'cell': cell.coordinate,
                'col': col_idx,
                'formula': str(cell.value)
            })
    if row_formulas:
        formulas_by_row[row_idx] = row_formulas

print("\nFormulas by row (first 30 rows with formulas):")
for row_idx in sorted(formulas_by_row.keys())[:30]:
    print(f"\nRow {row_idx}:")
    for f in formulas_by_row[row_idx]:
        print(f"  {f['cell']}: {f['formula']}")

# Try to understand the structure
print("\n" + "=" * 100)
print("STRUCTURE INFERENCE")
print("=" * 100)

# Read first column to understand row labels
first_col = []
for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row), min_col=1, max_col=1):
    for cell in row:
        first_col.append(cell.value)

print("\nFirst Column (Row Labels):")
for i, val in enumerate(first_col[:30], 1):
    if val:
        print(f"  Row {i}: {val}")

# Save detailed analysis
with open("temp_scripts/scenario_calculator_analysis.md", "w", encoding="utf-8") as f:
    f.write("# Scenario Calculator Analysis\n\n")
    f.write("## User Inputs (Yellow Cells)\n\n")
    for inp in yellow_inputs:
        f.write(f"- {inp['cell']}: {inp['value']}\n")
    f.write("\n## Formulas\n\n")
    for row_idx in sorted(formulas_by_row.keys())[:50]:
        f.write(f"### Row {row_idx}\n\n")
        for f in formulas_by_row[row_idx]:
            f.write(f"- {f['cell']}: `{f['formula']}`\n")
        f.write("\n")

print("\nDetailed analysis saved to temp_scripts/scenario_calculator_analysis.md")

