"""Detailed analysis of Scenario Calculator - extract ALL data with proper Hebrew encoding."""
import pandas as pd
from pathlib import Path
import openpyxl

excel_file = Path("דירה להשקעה עדכון  V3.2 8.24.xlsx")

wb = openpyxl.load_workbook(excel_file, data_only=False)
ws = wb["Scenarios Calculator"]

# Extract all row labels and values with proper encoding
output_lines = []
output_lines.append("# Scenario Calculator - Full Row Analysis\n\n")

output_lines.append("## Row-by-Row Analysis (Column A = Label, B/C/D = Values)\n\n")
output_lines.append("| Row | Hebrew Label | Value A | Value B | Value C | Formula (B) |\n")
output_lines.append("|-----|-------------|---------|---------|---------|-------------|\n")

for row_idx in range(1, min(80, ws.max_row + 1)):
    # Get cells
    label_cell = ws.cell(row=row_idx, column=1)  # Column A - label
    val_a = ws.cell(row=row_idx, column=2)  # Column B (Scenario A)
    val_b = ws.cell(row=row_idx, column=3)  # Column C (Scenario B) 
    val_c = ws.cell(row=row_idx, column=4)  # Column D (Scenario C)
    
    label = str(label_cell.value) if label_cell.value else ""
    
    # Get value or formula
    val_a_str = ""
    formula_a = ""
    if val_a.value is not None:
        if val_a.data_type == 'f':
            formula_a = str(val_a.value)
            val_a_str = "(formula)"
        else:
            val_a_str = str(val_a.value)
    
    val_b_str = str(val_b.value) if val_b.value else ""
    val_c_str = str(val_c.value) if val_c.value else ""
    
    # Check if yellow (user input)
    is_input = ""
    if val_a.fill and val_a.fill.patternType == 'solid':
        fill = val_a.fill.fgColor
        if fill and fill.rgb and ('FFFF' in str(fill.rgb).upper() or 'FFE' in str(fill.rgb).upper()):
            is_input = " **[INPUT]**"
    
    if label or val_a_str:
        output_lines.append(f"| {row_idx} | {label[:50]}{is_input} | {val_a_str[:15]} | {val_b_str[:15]} | {val_c_str[:15]} | {formula_a[:40]} |\n")

output_lines.append("\n\n## All Formulas in Column B (Scenario A)\n\n")

for row_idx in range(1, min(80, ws.max_row + 1)):
    cell = ws.cell(row=row_idx, column=2)
    label_cell = ws.cell(row=row_idx, column=1)
    
    if cell.data_type == 'f':
        label = str(label_cell.value) if label_cell.value else f"Row {row_idx}"
        output_lines.append(f"### Row {row_idx}: {label[:60]}\n")
        output_lines.append(f"```\n{cell.value}\n```\n\n")

# Save with UTF-8 encoding
with open("temp_scripts/scenario_full_analysis.md", "w", encoding="utf-8") as f:
    f.writelines(output_lines)

print("Analysis saved to temp_scripts/scenario_full_analysis.md")

# Also print key rows
print("\n" + "="*80)
print("KEY ROWS SUMMARY")
print("="*80)

key_rows = [2, 3, 4, 5, 6, 7, 8, 12, 14, 18, 19, 20, 21, 22, 25, 26, 27, 30, 31, 32, 33]
for row_idx in key_rows:
    label_cell = ws.cell(row=row_idx, column=1)
    val_cell = ws.cell(row=row_idx, column=2)
    
    label = str(label_cell.value) if label_cell.value else ""
    val = val_cell.value
    
    print(f"Row {row_idx:2d}: {label[:60]}")
    if val_cell.data_type == 'f':
        print(f"         Formula: {val}")
    else:
        print(f"         Value: {val}")
    print()

