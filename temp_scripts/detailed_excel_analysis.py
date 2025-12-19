"""Detailed Excel analysis with Hebrew column translation and logic understanding."""
import pandas as pd
from pathlib import Path
import openpyxl

excel_file = Path("דירה להשקעה עדכון  V3.2 8.24.xlsx")

print("=" * 100)
print("DETAILED EXCEL FILE ANALYSIS")
print("=" * 100)
print()

wb = openpyxl.load_workbook(excel_file, data_only=False)
sheet_names = wb.sheetnames

# Hebrew column name translations (based on context)
HEBREW_TRANSLATIONS = {
    "דירה": "Apartment",
    "עיר": "City",
    "שכונה": "Neighborhood", 
    "רחוב": "Street",
    "קישור": "Link/URL",
    "מחיר": "Price",
    "חדרים": "Rooms",
    "קומה": "Floor",
    "מ\"ר": "Square Meters",
    "מחיר למ\"ר": "Price per sqm",
    "תאריך": "Date",
    "דירה להשקעה": "Investment Apartment",
    "מחירון": "Price List",
    "תרחישים": "Scenarios",
    "מחשבון": "Calculator"
}

for idx, sheet_name in enumerate(sheet_names, 1):
    print("=" * 100)
    print(f"SHEET {idx}/{len(sheet_names)}: {sheet_name}")
    print("=" * 100)
    
    df = pd.read_excel(excel_file, sheet_name=sheet_name, engine='openpyxl', header=0)
    
    print(f"Dimensions: {df.shape[0]} rows × {df.shape[1]} columns")
    print()
    
    # Show first row to understand structure
    print("Column Headers:")
    for i, col in enumerate(df.columns, 1):
        print(f"  {i:2d}. {col}")
    print()
    
    # Show first few data rows
    print("First 2 Data Rows:")
    print(df.head(2).to_string())
    print()
    
    # Analyze formulas
    ws = wb[sheet_name]
    formula_examples = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=min(10, ws.max_row)), start=2):
        for cell in row:
            if cell.data_type == 'f' and len(formula_examples) < 10:
                formula_examples.append({
                    'cell': cell.coordinate,
                    'formula': str(cell.value),
                    'row': row_idx
                })
    
    if formula_examples:
        print("Sample Formulas:")
        for f in formula_examples[:10]:
            print(f"  {f['cell']} (row {f['row']}): {f['formula']}")
        print()
    
    # Try to understand sheet purpose
    print("PURPOSE ANALYSIS:")
    if "Data Table" in sheet_name:
        print("  → Main data table with apartment listings")
        print("  → Contains: URLs, location, price, size, calculated metrics")
        print("  → Has calculated fields: predicted price, ranks, scores")
    elif "Regression" in sheet_name:
        print("  → Statistical regression analysis output")
        print("  → Contains: R-squared, coefficients, significance tests")
        print("  → Purpose: Model validation and feature importance")
    elif "Scenarios" in sheet_name or "תרחישים" in sheet_name:
        print("  → Scenario calculator for investment analysis")
        print("  → Contains: Different investment scenarios (A, B, C)")
        print("  → Purpose: Calculate ROI, cash flow, etc. under different assumptions")
    elif "מחירון" in sheet_name:
        print("  → Price list/reference data")
        print("  → Contains: Historical prices, market data")
    else:
        print("  → Need to analyze further")
    print()

print("=" * 100)
print("KEY INSIGHTS")
print("=" * 100)
print("""
1. DATA TABLE: Main apartment listing data with calculated metrics
2. REGRESSION: Statistical model for price prediction
3. SCENARIOS CALCULATOR: Investment analysis tool
4. Other sheets: Supporting data and calculations

Main Functions Identified:
- Price prediction using regression model
- Investment scoring/ranking system
- Scenario analysis for ROI calculations
- Data processing and metric calculations
""")

